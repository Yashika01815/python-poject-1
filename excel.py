import openpyxl as xl
from openpyxl.chart import BarChart, Reference
wb = xl.load_workbook('transactions (1).xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(1, 3)
print(cell.value)
for row in range(2, sheet.max_row + 1):  # Start from row 2 since row 1 is the header
    print(row)
    
    # Fetch the original price from column 3 
    price = sheet.cell(row, 3).value
    # Calculate 5% and 10% increments
    increment_5 = price * 1.05
    increment_10 = price * 1.10
    # Write the new values into columns 4 and  5
    sheet.cell(row, 4).value = increment_5
    sheet.cell(row, 5).value = increment_10
sheet.cell(1, 4).value = 'Price + 5%'
sheet.cell(1, 5).value = 'Price + 10%'

#  (column 3, starting from row 2)
values = Reference(sheet,
                   min_row=2,
                   max_row=sheet.max_row,
                   min_col=3,
                   max_col=3)
# Create a bar chart
chart = BarChart()
chart.title = "Product ID Distribution"
chart.x_axis.title = "Transaction"
chart.y_axis.title = "Product ID"
chart.add_data(values)
# Add the chart to the sheet at position E2
sheet.add_chart(chart, 'E2')
# workbook with changes
wb.save('transaction.xlsx')

